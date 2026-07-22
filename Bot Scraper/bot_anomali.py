"""
Rekap Anomali Pendataan (Usaha & Keluarga) Pertanggal
Pipeline: login headless -> scrap anomali usaha & keluarga per kabupaten
          -> gabung usaha+keluarga -> pisah resolved/belum resolved
          -> gabung dgn rekap kemarin (outer join) -> tandai resolved
          -> simpan Excel (sheet Rekap + Ringkasan) & upsert ke SQLite.

Ini adalah gabungan dari:
- pola login headless + retry per kabupaten (dari script rekap progres pendataan)
- pola rekap-pertanggal / checklist resolved (dari script rekap anomali lama)
"""

import json
import os
import random
import re
import sqlite3
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
import pyotp
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Font
from playwright.sync_api import sync_playwright
from playwright_stealth import Stealth

# Cari .env di folder yang sama dengan file script ini (sama seperti script
# rekap progres pendataan), supaya tidak tergantung working directory.
ENV_PATH = Path(__file__).resolve().parent / ".env"
load_dotenv(dotenv_path=ENV_PATH)

# ── Konfigurasi umum ─────────────────────────────────────────────────────

FORMAT_TANGGAL = "%d-%m-%Y"
FORMAT_TANGGAL_JAM = "%d-%m-%Y_%H-%M-%S"
TANGGAL_HARI_INI = datetime.now().strftime(FORMAT_TANGGAL)
TANGGAL_JAM = datetime.now().strftime(FORMAT_TANGGAL_JAM)

URL_DASHBOARD = "https://dashboard-se2026.apps.bps.go.id"
URL_API_ANOMALI = "https://dashboard-se2026.apps.bps.go.id/api/mikro/anomali-case-kab"
MAX_RETRY_PER_KAB = 5

# Kredensial dari .env, sama seperti script rekap progres pendataan.
DASHBOARD_USERNAME = os.environ.get("DASHBOARD_USERNAME")
DASHBOARD_PASSWORD = os.environ.get("DASHBOARD_PASSWORD")
DASHBOARD_OTP_SECRET = os.environ.get("DASHBOARD_OTP_SECRET")

_KREDENSIAL = {
    "DASHBOARD_USERNAME": DASHBOARD_USERNAME,
    "DASHBOARD_PASSWORD": DASHBOARD_PASSWORD,
    "DASHBOARD_OTP_SECRET": DASHBOARD_OTP_SECRET,
}
_KOSONG = [k for k, v in _KREDENSIAL.items() if not v]

if _KOSONG:
    raise RuntimeError(
        f"Variabel berikut kosong/tidak ditemukan: {', '.join(_KOSONG)}.\n"
        f"Dicari dari file: {ENV_PATH} (ada: {ENV_PATH.exists()}).\n"
        "Pastikan file .env ada persis di folder yang sama dengan script ini "
        "dan berisi baris seperti:\n"
        "  DASHBOARD_USERNAME=muh.prayitno\n"
        "  DASHBOARD_PASSWORD=Prayitno_26\n"
        "  DASHBOARD_OTP_SECRET=MRVGI53EMIYTKMTHOZ3W64RXJU2G6R3H"
    )

SELECTOR_USERNAME = "xpath=//*[@id='username']"
SELECTOR_PASSWORD = "xpath=//*[@id='password']"
SELECTOR_OTP = "xpath=//*[@id='otp']"
SELECTOR_TOMBOL_LOGIN = "xpath=//*[@id='v-0']/button"

KODE_KAB_LIST = [
    "1601",
    "1602",
    "1603",
    "1604",
    "1605",
    "1606",
    "1607",
    "1608",
    "1609",
    "1610",
    "1611",
    "1612",
    "1613",
    "1671",
    "1672",
    "1673",
    "1674",
]

# Jeda antar-request, sama seperti script rekap progres pendataan, supaya
# pola trafik tidak terlihat seperti bot.
DELAY_ANTAR_KABUPATEN_DETIK = (15, 30)

FOLDER_BACKUP_USAHA = Path("../scrap_anomali_usaha")
FOLDER_BACKUP_KELUARGA = Path("../scrap_anomali_keluarga")
FOLDER_REKAP = Path("../rekap_anomali_pertanggal")
FOLDER_OUTPUT_DB = Path("../SQLLITE")
DB_PATH = FOLDER_OUTPUT_DB / "rekap_anomali_pendataan.db"
TABLE_NAME = "rekap_anomali_pendataan"

# Parameter API per jenis anomali, persis dari 2 script asal.
PARAM_ANOMALI_USAHA = {
    "indikator": "128,129,130,131,132,133,134,135",
    "sudah_indikator": "40,41,42,43,44,45,46,46",
    "type": "usaha",
    "anomali_no": "1,2,3,4,5,6,7,8",
}
PARAM_ANOMALI_KELUARGA = {
    "indikator": "136,137,139,140,141,142,144",
    "sudah_indikator": "47,48,50,51,52,53",
    "type": "keluarga",
    "anomali_no": "1,2,3,4,5,6,7",
}

KEY_COLS = [
    "assignment_id",
    "kode_kabupaten",
    "level_6_code",
    "nama_tercantum",
    "source_type",
    "anomali_no",
]
EXTRA_COLS = ["extra_columns"]
SELECT_COLS = KEY_COLS + ["is_resolved"] + EXTRA_COLS

TANDA_CENTANG = "\u2713"
TANDA_KONDISI_LAPANGAN = "Kondisi Lapangan"
MAX_LEBAR_KOLOM = 40

LINK_BASE_URL = (
    "https://fasih-sm.bps.go.id/app/assignment/fd68e454-ba45-4b85-8205-f3bf777ded24/"
)
LINK_COL = "hyperlink"


# ── Bagian 1: Login headless (sama seperti script rekap progres pendataan) ──


def isi_username_password(page) -> None:
    """Isi form username & password, lalu submit dengan menekan Enter."""
    page.wait_for_selector(SELECTOR_USERNAME, state="visible", timeout=15000)
    page.fill(SELECTOR_USERNAME, DASHBOARD_USERNAME)
    page.fill(SELECTOR_PASSWORD, DASHBOARD_PASSWORD)
    page.press(SELECTOR_PASSWORD, "Enter")
    print("Username & password terisi, form disubmit.")


def isi_otp(page, timeout_ms: int = 20000) -> None:
    """Isi kode OTP dari TOTP generator, lalu submit dengan Enter.
    Kode OTP baru di-generate saat elemen sudah muncul, supaya tidak
    keburu expired kalau ada delay sebelumnya."""
    page.wait_for_selector(SELECTOR_OTP, state="visible", timeout=timeout_ms)

    totp = pyotp.TOTP(DASHBOARD_OTP_SECRET)
    otp_code = totp.now()

    page.fill(SELECTOR_OTP, otp_code)
    page.press(SELECTOR_OTP, "Enter")
    print(f"OTP ({otp_code}) terisi, form disubmit.")


def klik_tombol_login(page, max_percobaan: int = 5, timeout_ms: int = 15000) -> bool:
    """Klik tombol login, verifikasi ada perubahan nyata di halaman.
    Kalau form username langsung muncul (tanpa captcha), sekalian isi
    username/password/OTP di sini.

    Return True kalau proses login (username+password+OTP) sudah selesai
    ditangani otomatis di dalam fungsi ini, False kalau belum (misal karena
    masih perlu captcha manual dulu, atau tombol login tidak ditemukan sama
    sekali karena sudah login lewat session_state)."""
    try:
        page.wait_for_selector(
            SELECTOR_TOMBOL_LOGIN, timeout=timeout_ms, state="visible"
        )
    except Exception as e:
        print(f"Tombol login tidak ditemukan (mungkin sudah login): {e}")
        return False

    url_sebelum = page.url

    for percobaan in range(1, max_percobaan + 1):
        try:
            time.sleep(random.uniform(1, 5))
            page.locator(SELECTOR_TOMBOL_LOGIN).click(timeout=5000)
        except Exception as e:
            print(f"  Percobaan {percobaan} - klik gagal: {e}")
            page.wait_for_timeout(1500)
            continue

        page.wait_for_timeout(2000)

        url_sesudah = page.url
        ada_perubahan_url = url_sesudah != url_sebelum
        ada_captcha = (
            page.locator("iframe[src*='captcha'], .captcha, #captcha").count() > 0
        )
        ada_form_username = page.locator(SELECTOR_USERNAME).count() > 0

        if ada_perubahan_url or ada_captcha or ada_form_username:
            print(
                f"Tombol login berhasil diklik (percobaan {percobaan}), halaman berubah."
            )

            if ada_form_username:
                isi_username_password(page)
                isi_otp(page)
                return True

            print("Form username belum muncul (kemungkinan perlu captcha manual dulu).")
            return False

        print(
            f"  Percobaan {percobaan} - klik terkirim tapi belum ada perubahan, coba lagi..."
        )

    print(
        "Tombol login diklik tapi tidak terdeteksi ada perubahan setelah beberapa percobaan."
    )
    return False


def login_dashboard(page) -> None:
    """Bungkus alur login lengkap: klik tombol login -> (captcha manual kalau
    perlu) -> isi username/password/OTP."""
    page.goto(URL_DASHBOARD)
    login_selesai_otomatis = klik_tombol_login(page)

    if not login_selesai_otomatis:
        input("Selesaikan captcha dulu, lalu tekan ENTER...")
        if page.locator(SELECTOR_USERNAME).count() > 0:
            isi_username_password(page)
            isi_otp(page)


# ── Bagian 2: Scraping anomali per kabupaten ────────────────────────────


def minta_data_anomali(context, page, kab: str, params_dasar: dict) -> list | None:
    """Minta data anomali satu kabupaten ke API. Retry otomatis kalau:
    - request error (koneksi putus, timeout, dll)
    - status bukan 200 (misal 401 = sesi login expired, 400 = bad request)
    - kena deteksi bot dari situs (tunggu 120 detik cooldown, lalu coba lagi)

    Semua retry dibatasi MAX_RETRY_PER_KAB -- kalau habis, kabupaten ini
    dilewati dan lanjut ke kabupaten berikutnya."""
    params = {**params_dasar, "kode_kabupaten": kab}

    for percobaan in range(1, MAX_RETRY_PER_KAB + 1):
        try:
            response = context.request.get(
                URL_API_ANOMALI, params=params, timeout=120000
            )
        except Exception as e:
            print(f"  [{kab}] percobaan {percobaan} - Request error: {e}")
            time.sleep(5)
            continue

        content_type = response.headers.get("content-type", "")
        print(
            f"  [{kab}] percobaan {percobaan} - Status: {response.status} - "
            f"Content-Type: {content_type}"
        )

        if response.status == 200 and "application/json" in content_type:
            return response.json()

        cuplikan = response.text()[:500]

        if "Bot Detected" in cuplikan or "terdeteksi sebagai bot" in cuplikan:
            print(f"  [{kab}] Kena deteksi bot. Menunggu 120 detik untuk cooldown...")
            time.sleep(120)
            page.goto(URL_DASHBOARD)
            continue

        if response.status in (401, 403):
            print(
                f"  [{kab}] Status {response.status} - sesi login kemungkinan expired."
            )
            print("  Cuplikan response:", cuplikan[:300])
            time.sleep(5)
            continue

        if response.status >= 400:
            print(f"  [{kab}] Status {response.status} - request gagal, coba lagi...")
            print("  Cuplikan response:", cuplikan[:300])
            time.sleep(5)
            continue

        print(f"  [{kab}] Response bukan JSON, sepertinya perlu captcha ulang.")
        print("  Cuplikan response:", cuplikan[:300])
        page.goto(URL_DASHBOARD)
        input(f"  Selesaikan captcha untuk lanjut scrap Kab {kab}, lalu tekan ENTER...")

    print(f"  [{kab}] GAGAL setelah {MAX_RETRY_PER_KAB} percobaan, dilewati.")
    return None


def scrap_anomali(context, page, params_dasar: dict, label: str) -> pd.DataFrame:
    """Scrap semua kabupaten untuk satu jenis anomali (usaha atau keluarga).
    Kalau satu kabupaten gagal terus, kabupaten itu dilewati dan proses
    lanjut ke kabupaten berikutnya."""
    semua_df = []
    kab_gagal = []

    for kab in KODE_KAB_LIST:
        print(f"[{label}] Proses Kab {kab}")
        data = minta_data_anomali(context, page, kab, params_dasar)
        if data is not None:
            df = pd.DataFrame(data)
            df["kode_kabupaten"] = kab
            semua_df.append(df)
            print(f"  -> {len(df)} baris")
        else:
            kab_gagal.append(kab)

        jeda = random.uniform(*DELAY_ANTAR_KABUPATEN_DETIK)
        print(f"  Jeda {jeda:.1f} detik sebelum kabupaten berikutnya...")
        time.sleep(jeda)

    if kab_gagal:
        print(
            f"\nPERINGATAN [{label}]: {len(kab_gagal)} kabupaten GAGAL di-scrap "
            f"dan dilewati: {kab_gagal}\n"
        )

    if not semua_df:
        raise RuntimeError(f"Tidak ada data anomali {label} yang berhasil diambil.")

    return pd.concat(semua_df, ignore_index=True)


def simpan_backup_excel(df: pd.DataFrame, folder: Path, prefix: str) -> Path:
    """Simpan hasil scraping mentah sebagai backup, sebelum diproses lebih lanjut."""
    folder.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path_output = folder / f"{prefix}_{timestamp}.xlsx"
    df.to_excel(path_output, index=False)
    print("Backup tersimpan di:", path_output)
    return path_output


def jalankan_scraping_anomali() -> tuple[pd.DataFrame, pd.DataFrame]:
    """Login headless sekali, lalu scrap anomali usaha & keluarga dalam satu
    sesi browser. Mengembalikan (df_usaha, df_keluarga) mentah."""
    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True, args=["--disable-blink-features=AutomationControlled"]
        )
        context = browser.new_context(
            accept_downloads=True,
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1366, "height": 768},
            locale="id-ID",
            timezone_id="Asia/Jakarta",
        )
        page = context.new_page()
        Stealth().apply_stealth_sync(page)

        login_dashboard(page)

        print("=== Scraping anomali USAHA ===")
        df_usaha = scrap_anomali(context, page, PARAM_ANOMALI_USAHA, "usaha")

        print("=== Scraping anomali KELUARGA ===")
        df_keluarga = scrap_anomali(context, page, PARAM_ANOMALI_KELUARGA, "keluarga")

        context.close()
        browser.close()

    simpan_backup_excel(df_usaha, FOLDER_BACKUP_USAHA, "anomali_usaha_sumsel")
    simpan_backup_excel(df_keluarga, FOLDER_BACKUP_KELUARGA, "anomali_keluarga_sumsel")

    return df_usaha, df_keluarga


# ── Bagian 3: Proses rekap pertanggal (checklist resolved) ─────────────

def _jsonify_dict_list(nilai):
    """API kadang mengembalikan extra_columns sebagai dict/list per baris.
    SQLite (dan sebagian proses merge/outer-join) tidak bisa menangani objek
    Python langsung, jadi diubah jadi string JSON supaya aman disimpan &
    dibandingkan sebagai teks biasa."""
    if isinstance(nilai, (dict, list)):
        return json.dumps(nilai, ensure_ascii=False)
    return nilai


def _normalisasi_kolom_key(df: pd.DataFrame) -> pd.DataFrame:
    """Pastikan semua kolom KEY_COLS selalu bertipe string sebelum dipakai
    untuk merge. Data anomali baru berasal dari JSON API (tipe kolom
    mengikuti apa pun yang dikirim server, bisa int atau str), sedangkan
    rekap lama dibaca dari Excel. Kalau tipenya beda (object vs int64),
    pd.merge akan error 'You are trying to merge on object and int64
    columns'. Disamakan ke string + strip supaya aman."""
    df = df.copy()
    for kolom in KEY_COLS:
        if kolom in df.columns:
            df[kolom] = df[kolom].astype(str).str.strip()
    return df


def siapkan_raw_anomali(
    df_usaha: pd.DataFrame, df_keluarga: pd.DataFrame
) -> pd.DataFrame:
    """Gabungkan data anomali usaha & keluarga, ambil kolom relevan saja.
    Kolom extra_columns langsung di-serialisasi ke JSON string di sini,
    supaya konsisten di semua tahap berikutnya (merge, Excel, SQLite).
    Kolom key juga langsung dinormalisasi ke string di sini, supaya
    is_resolved/not_resolved yang diturunkan darinya ikut konsisten."""
    anomali_usaha = df_usaha[SELECT_COLS].copy()
    anomali_keluarga = df_keluarga[SELECT_COLS].copy()
    gabungan = pd.concat([anomali_usaha, anomali_keluarga], ignore_index=True)

    for kolom in EXTRA_COLS:
        gabungan[kolom] = gabungan[kolom].map(_jsonify_dict_list)

    gabungan = _normalisasi_kolom_key(gabungan)

    return gabungan


def pisahkan_status_resolved(
    raw_anomali: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Pisahkan anomali menjadi (is_resolved, not_resolved).

    Catatan: kolom is_resolved berisi string 'True'/'False', bukan boolean,
    karena df dibaca lewat JSON API lalu jadi kolom object/str."""
    is_resolved = raw_anomali[raw_anomali["is_resolved"].astype(str) == "True"]
    not_resolved = raw_anomali[raw_anomali["is_resolved"].astype(str) == "False"]
    return is_resolved, not_resolved


def muat_rekap_terakhir(folder_rekap: Path) -> pd.DataFrame | None:
    """Cari & muat file rekap terakhir berdasarkan waktu modifikasi.
    Return None kalau belum ada file rekap sama sekali."""
    existing_files = list(folder_rekap.glob("rekap_anomali_*.xlsx"))
    if not existing_files:
        print("Belum ada file rekap sebelumnya, akan membuat baru.")
        return None

    file_terakhir = max(existing_files, key=lambda f: f.stat().st_mtime)
    print("File rekap terakhir ditemukan:", file_terakhir)
    df_lama = pd.read_excel(file_terakhir, sheet_name="Rekap", dtype=str)
    return _normalisasi_kolom_key(df_lama)


def gabungkan_dengan_rekap_lama(
    df_lama: pd.DataFrame | None,
    not_resolved: pd.DataFrame,
    tanggal_hari_ini: str,
) -> pd.DataFrame:
    """Outer join rekap lama dengan anomali yang belum resolved hari ini.

    Baris yang belum resolved diberi tanda centang pada kolom tanggal
    hari ini. Baris lama yang tidak muncul lagi akan tetap ada (kolom
    tanggal hari ini menjadi kosong/NaN untuk baris tsb)."""
    raw_anomali_baru = not_resolved.drop(columns=["is_resolved"]).copy()
    raw_anomali_baru[tanggal_hari_ini] = TANDA_CENTANG

    if df_lama is None:
        print(
            f"Belum ada data sebelumnya, semua {len(raw_anomali_baru)} baris dianggap baru."
        )
        return raw_anomali_baru

    rekap = pd.merge(
        df_lama,
        raw_anomali_baru,
        on=KEY_COLS,
        how="outer",
        suffixes=("", "_baru"),
        indicator=True,
    )
    for col in EXTRA_COLS:
        col_baru = f"{col}_baru"
        if col_baru in rekap.columns:
            rekap[col] = rekap[col_baru].combine_first(rekap[col])
            rekap = rekap.drop(columns=[col_baru])
    return rekap.drop(columns=["_merge"])


def tandai_resolved(
    rekap: pd.DataFrame,
    is_resolved: pd.DataFrame,
    tanggal_hari_ini: str,
) -> pd.DataFrame:
    """Tandai baris yang key-nya cocok dengan anomali is_resolved.
    Hanya baris yang match yang diubah."""
    rekap = pd.merge(
        rekap,
        is_resolved[KEY_COLS],
        on=KEY_COLS,
        how="left",
        indicator="_merge_resolved",
    )
    rekap.loc[rekap["_merge_resolved"] == "both", tanggal_hari_ini] = (
        TANDA_KONDISI_LAPANGAN
    )
    return rekap.drop(columns=["_merge_resolved"])


def _is_kolom_tanggal(col: str) -> bool:
    try:
        datetime.strptime(col, FORMAT_TANGGAL)
        return True
    except (ValueError, TypeError):
        return False


def cari_tanggal_sebelumnya(rekap: pd.DataFrame, tanggal_hari_ini: str) -> str | None:
    """Cari kolom tanggal terakhir sebelum tanggal_hari_ini (untuk pembanding)."""
    kolom_tanggal = [c for c in rekap.columns if _is_kolom_tanggal(c)]
    kolom_tanggal_sorted = sorted(
        kolom_tanggal, key=lambda c: datetime.strptime(c, FORMAT_TANGGAL)
    )
    kolom_sebelum_hari_ini = [c for c in kolom_tanggal_sorted if c != tanggal_hari_ini]

    print("Kolom tanggal terdeteksi:", kolom_tanggal_sorted)
    return kolom_sebelum_hari_ini[-1] if kolom_sebelum_hari_ini else None


def buat_ringkasan(
    rekap: pd.DataFrame,
    df_lama: pd.DataFrame | None,
    raw_anomali_baru_count: int,
    tanggal_hari_ini: str,
    tanggal_sebelumnya: str | None,
) -> pd.DataFrame:
    """Susun tabel ringkasan: jumlah data lama/baru & perubahan status centang."""
    jumlah_data_sebelumnya = len(df_lama) if df_lama is not None else 0

    if tanggal_sebelumnya is not None:
        kolom_now = rekap[tanggal_hari_ini].fillna("")
        kolom_prev = rekap[tanggal_sebelumnya].fillna("")

        was_centang = kolom_prev == TANDA_CENTANG
        now_centang = kolom_now == TANDA_CENTANG

        jumlah_berubah = (was_centang & ~now_centang).sum()
        jumlah_belum_berubah = (was_centang & now_centang).sum()
    else:
        jumlah_berubah = None
        jumlah_belum_berubah = None
        print(
            "Belum ada kolom tanggal sebelumnya untuk dibandingkan (rekap pertama kali)."
        )

    return pd.DataFrame(
        {
            "Keterangan": [
                "Data sebelumnya",
                "Data baru masuk",
                f"Perubahan (centang -> blank/{TANDA_KONDISI_LAPANGAN}), "
                f"{tanggal_sebelumnya} -> {tanggal_hari_ini}",
                f"Belum berubah (tetap centang), {tanggal_sebelumnya} -> {tanggal_hari_ini}",
                "Total setelah digabung",
            ],
            "Jumlah": [
                jumlah_data_sebelumnya,
                raw_anomali_baru_count,
                jumlah_berubah,
                jumlah_belum_berubah,
                len(rekap),
            ],
        }
    )


def tambah_kolom_link(rekap: pd.DataFrame) -> pd.DataFrame:
    """Tambah kolom link = LINK_BASE_URL + assignment_id."""
    rekap = rekap.copy()
    rekap[LINK_COL] = LINK_BASE_URL + rekap["assignment_id"].astype(str)
    return rekap


def urutkan_kolom_rekap(rekap: pd.DataFrame) -> pd.DataFrame:
    """Susun ulang urutan kolom: key -> extra_columns -> hyperlink -> tanggal (kronologis)."""
    kolom_tanggal = [c for c in rekap.columns if _is_kolom_tanggal(c)]
    kolom_tanggal_sorted = sorted(
        kolom_tanggal, key=lambda c: datetime.strptime(c, FORMAT_TANGGAL)
    )

    kolom_urut = KEY_COLS + EXTRA_COLS + [LINK_COL] + kolom_tanggal_sorted
    kolom_sisa = [c for c in rekap.columns if c not in kolom_urut]
    return rekap[kolom_urut + kolom_sisa]


# ── Bagian 4: Export ke Excel ────────────────────────────────────────────


def _jadikan_hyperlink(
    path_file: Path, sheet_name: str = "Rekap", kolom_link: str = LINK_COL
) -> None:
    """Ubah teks URL di kolom link menjadi hyperlink asli yang bisa diklik."""
    wb = load_workbook(path_file)
    ws = wb[sheet_name]

    header = [cell.value for cell in ws[1]]
    if kolom_link not in header:
        wb.close()
        return
    col_idx = header.index(kolom_link) + 1

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value:
            cell.hyperlink = cell.value
            cell.font = Font(color="0563C1", underline="single")

    wb.save(path_file)


def _atur_lebar_kolom(path_file: Path, max_lebar: int = MAX_LEBAR_KOLOM) -> None:
    """Sesuaikan lebar kolom tiap sheet otomatis berdasarkan isi terpanjang."""
    wb = load_workbook(path_file)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for col_cells in ws.columns:
            max_len = max(
                (len(str(cell.value)) if cell.value is not None else 0)
                for cell in col_cells
            )
            col_letter = col_cells[0].column_letter
            ws.column_dimensions[col_letter].width = min(max_len + 2, max_lebar)
    wb.save(path_file)


def simpan_ke_excel(
    rekap: pd.DataFrame, ringkasan: pd.DataFrame, path_file: Path
) -> None:
    with pd.ExcelWriter(path_file, engine="openpyxl") as writer:
        rekap.to_excel(writer, sheet_name="Rekap", index=False)
        ringkasan.to_excel(writer, sheet_name="Ringkasan", index=False)

    _jadikan_hyperlink(path_file)
    _atur_lebar_kolom(path_file)
    print("Excel tersimpan di:", path_file)


# ── Bagian 5: Upsert ke SQLite ───────────────────────────────────────────
# Beda dengan rekap progres pendataan (key tunggal id_wilayah), di sini
# key-nya gabungan 6 kolom (KEY_COLS), jadi dipakai UNIQUE INDEX komposit.
# Kolom tanggal baru (mis. "22-07-2026") ditambahkan otomatis tiap ada
# tanggal baru, sama seperti kolom indikator baru di script satunya.


def pastikan_tabel_dan_kolom(df: pd.DataFrame, db_path: Path, table: str) -> None:
    """Buat tabel + unique index komposit kalau belum ada; tambah kolom baru
    kalau df punya kolom yang belum dikenal tabel (mis. kolom tanggal baru)."""
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    cursor.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table,)
    )
    tabel_ada = cursor.fetchone()

    if not tabel_ada:
        df.head(0).to_sql(table, conn, if_exists="replace", index=False)
        kolom_key = ", ".join(f'"{k}"' for k in KEY_COLS)
        cursor.execute(
            f'CREATE UNIQUE INDEX IF NOT EXISTS idx_key_anomali ON "{table}" ({kolom_key})'
        )
    else:
        cursor.execute(f'PRAGMA table_info("{table}")')
        kolom_ada = {row[1] for row in cursor.fetchall()}
        for kolom in df.columns:
            if kolom not in kolom_ada:
                cursor.execute(f'ALTER TABLE "{table}" ADD COLUMN "{kolom}" TEXT')
                print(f"Kolom '{kolom}' ditambahkan ke tabel {table}")

    conn.commit()
    conn.close()


def upsert_ke_sqlite(
    df: pd.DataFrame, tanggal_jam: str, db_path: Path = DB_PATH, table: str = TABLE_NAME
) -> None:
    """Upsert (insert atau replace) berdasarkan key komposit KEY_COLS.
    Karena df rekap sudah berisi seluruh histori kolom tanggal (hasil outer
    join kumulatif), tiap baris di-replace utuh dengan versi terbaru."""
    df = df.copy()
    df["last_update"] = tanggal_jam

    FOLDER_OUTPUT_DB.mkdir(parents=True, exist_ok=True)
    pastikan_tabel_dan_kolom(df, db_path, table)

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    kolom_str = ", ".join(f'"{k}"' for k in df.columns)
    placeholder = ", ".join(["?"] * len(df.columns))
    data = df.where(pd.notnull(df), None).values.tolist()

    # Pengaman tambahan: kalau ada sel berisi dict/list (mis. extra_columns
    # yang lolos dari serialisasi di siapkan_raw_anomali), ubah ke JSON
    # string di sini supaya sqlite3 tidak error "type 'dict' is not supported".
    data = [
        [
            json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v
            for v in baris
        ]
        for baris in data
    ]

    cursor.executemany(
        f'INSERT OR REPLACE INTO "{table}" ({kolom_str}) VALUES ({placeholder})', data
    )
    conn.commit()
    conn.close()

    print(f"{len(df)} baris berhasil di-upsert ke {table} pada {tanggal_jam}")


# ── Main ──────────────────────────────────────────────────────────────────


def main() -> None:
    FOLDER_REKAP.mkdir(parents=True, exist_ok=True)

    print("=== Proses Login & Scraping Anomali (headless) ===")
    df_usaha, df_keluarga = jalankan_scraping_anomali()
    # df_keluarga = pd.read_excel(
    #     "../scrap_anomali_keluarga/anomali_keluarga_sumsel_20260722_122744.xlsx"
    # )
    # df_usaha = pd.read_excel(
    #     "../scrap_anomali_usaha/anomali_usaha_sumsel_20260722_122720.xlsx"
    # )
    print("=== Proses Rekap Pertanggal ===")
    raw_anomali = siapkan_raw_anomali(df_usaha, df_keluarga)
    is_resolved, not_resolved = pisahkan_status_resolved(raw_anomali)
    print(f"Anomali resolved: {len(is_resolved)} | belum resolved: {len(not_resolved)}")

    df_lama = muat_rekap_terakhir(FOLDER_REKAP)

    rekap = gabungkan_dengan_rekap_lama(df_lama, not_resolved, TANGGAL_HARI_INI)
    rekap = tandai_resolved(rekap, is_resolved, TANGGAL_HARI_INI)

    tanggal_sebelumnya = cari_tanggal_sebelumnya(rekap, TANGGAL_HARI_INI)
    ringkasan = buat_ringkasan(
        rekap,
        df_lama,
        raw_anomali_baru_count=len(not_resolved),
        tanggal_hari_ini=TANGGAL_HARI_INI,
        tanggal_sebelumnya=tanggal_sebelumnya,
    )
    print(ringkasan.to_string(index=False))

    rekap = tambah_kolom_link(rekap)
    rekap = urutkan_kolom_rekap(rekap)

    nama_file_baru = FOLDER_REKAP / f"rekap_anomali_{TANGGAL_JAM}.xlsx"
    simpan_ke_excel(rekap, ringkasan, nama_file_baru)
    
    print("=== Proses Upsert ke SQLite ===")
    upsert_ke_sqlite(rekap, TANGGAL_JAM)

    print("Selesai.")


if __name__ == "__main__":
    main()
